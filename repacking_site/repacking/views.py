import csv

import openpyxl
from django.contrib.auth.decorators import login_required, permission_required
from django.http import Http404, HttpResponseRedirect, HttpResponse
from django.shortcuts import render, redirect
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Fill, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet import drawing

from logs.models import Log
from repacking_site.methods import filtered_records
from .filters import *
from .forms import *
from .models import *

repack_start_key = 'repack_start'
paused_repack_last_start_key = 'paused_repack_last_start'
repack_last_start_key = 'repack_last_start'
repack_duration_key = 'repack_duration'
repack_time_format = '%Y-%m-%dT%H:%M:%S'


@login_required
def index(request):
    cancel_sessions(request)
    return render(request, 'index.html')


@login_required
def repacking(request, sku_code, destination, idp_code, operators):
    standard = RepackingStandard.get_standard(sku_code, destination)
    if standard is None:
        raise Http404("Standard does not exist")

    if request.session.get(repack_start_key, None) is None:
        request.session[repack_start_key] = datetime.now().strftime(repack_time_format)
        request.session[repack_duration_key] = 0
    if request.session.get(repack_last_start_key, None) is None:
        request.session[repack_last_start_key] = datetime.now().strftime(repack_time_format)
    list_of_operator_names = []
    for operator in operators.split(','):
        list_of_operator_names.append(User.objects.get(barcode=operator).username)

    return render(request, 'repacking/repack.html', {'standard': standard, 'idp': idp_code, 'operators': operators,
                                                     'duration': int(request.session[repack_duration_key]),
                                                     repack_last_start_key: request.session[repack_last_start_key],
                                                     'list_of_operator_names': list_of_operator_names})


@login_required
def detail(request, sku_code, destination):
    cancel_sessions(request)
    standard = RepackingStandard.get_standard(sku_code, destination)
    if standard is None:
        raise Http404("Standard does not exist")
    return render(request, 'repacking/detail.html', {'standard': standard})


@login_required
def export(request, sku_code, destination):
    cancel_sessions(request)
    standard = RepackingStandard.get_standard(sku_code, destination)
    if standard is None:
        raise Http404("Standard does not exist")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={sku_code}.xlsx'
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = f'{sku_code}'

    orange_fill = PatternFill(
        start_color='f3d39b',
        end_color='f3d39b',
        fill_type='solid',
    )

    blue_fill = PatternFill(
        start_color='b3e4f3',
        end_color='b3e4f3',
        fill_type='solid',
    )

    green_fill = PatternFill(
        start_color='b9e5a0',
        end_color='b9e5a0',
        fill_type='solid',
    )

    side = Side(border_style='thin')

    border = Border(left=side, right=side, top=side, bottom=side)

    worksheet.cell(row=1, column=1).fill = orange_fill
    worksheet.cell(row=1, column=1).value = f'Referencia:'
    # worksheet.cell(row=1, column=1).border = border
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    worksheet.cell(row=1, column=3).fill = orange_fill
    worksheet.cell(row=1, column=3).value = f'Cofor:'
    # worksheet.cell(row=1, column=3).border = border
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
    worksheet.cell(row=1, column=5).fill = orange_fill
    worksheet.cell(row=1, column=5).value = f'Dodávateľ:'
    # worksheet.cell(row=1, column=5).border = border
    worksheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
    worksheet.cell(row=1, column=7).fill = orange_fill
    worksheet.cell(row=1, column=7).value = f'Destinácia:'
    # worksheet.cell(row=1, column=7).border = border
    worksheet.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)

    worksheet.row_dimensions[1].height = 35

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    worksheet.cell(row=2, column=1).value = f'{standard.SKU}'
    # worksheet.cell(row=2, column=1).border = border
    worksheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
    worksheet.cell(row=2, column=3).value = f'{standard.COFOR}'
    # worksheet.cell(row=2, column=3).border = border
    worksheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)
    worksheet.cell(row=2, column=5).value = f'{standard.supplier}'
    # worksheet.cell(row=2, column=5).border = border
    worksheet.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
    worksheet.cell(row=2, column=7).value = f'{standard.destination}'
    # worksheet.cell(row=2, column=7).border = border

    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    worksheet.cell(row=3, column=1).value = f'Balenie na príjme:'
    worksheet.row_dimensions[4].height = 100
    for i, image in enumerate(standard.input_photos.all()):
        img = Image(settings.MEDIA_ROOT + image.photo.url)
        ratio = img.height / 100
        img.height = img.height / ratio
        img.width = img.width / ratio
        worksheet.add_image(img, f'{get_column_letter(i + 1)}4')

    worksheet.merge_cells(start_row=3, start_column=5, end_row=3, end_column=8)
    worksheet.cell(row=3, column=5).value = f'Balenie na expedícii:'
    worksheet.row_dimensions[4].height = 100
    for i, image in enumerate(standard.output_photos.all()):
        img = Image(settings.MEDIA_ROOT + image.photo.url)
        ratio = img.height / 100
        img.height = img.height / ratio
        img.width = img.width / ratio
        worksheet.add_image(img, f'{get_column_letter(i + 5)}4')

    worksheet.cell(row=5, column=1).value = f'OPP:'
    worksheet.row_dimensions[5].height = 75
    for i, image in enumerate(standard.tools.all()):
        img = Image(settings.MEDIA_ROOT + image.photo.url)
        ratio = img.height/75
        img.height = img.height/ratio
        img.width = img.width/ratio
        worksheet.add_image(img, f'{get_column_letter(i+2)}5')

    worksheet.cell(row=6, column=1).value = f'Typ balenia:'
    worksheet.cell(row=6, column=1).fill = blue_fill
    # worksheet.cell(row=6, column=1).border = border
    worksheet.cell(row=6, column=2).value = f'{standard.input_type_of_package}'
    # worksheet.cell(row=6, column=2).border = border
    worksheet.cell(row=6, column=3).fill = blue_fill
    worksheet.cell(row=6, column=3).value = f'Počet boxov na palete:'
    # worksheet.cell(row=6, column=3).border = border
    worksheet.cell(row=6, column=4).value = f'{standard.input_count_of_boxes_on_pallet}'
    # worksheet.cell(row=6, column=4).border = border
    worksheet.cell(row=6, column=5).fill = green_fill
    worksheet.cell(row=6, column=5).value = f'Typ balenia:'
    # worksheet.cell(row=6, column=5).border = border
    worksheet.cell(row=6, column=6).value = f'{standard.output_type_of_package}'
    # worksheet.cell(row=6, column=6).border = border
    worksheet.cell(row=6, column=7).fill = green_fill
    worksheet.cell(row=6, column=7).value = f'Počet boxov na palete:'
    # worksheet.cell(row=6, column=7).border = border
    worksheet.cell(row=6, column=8).value = f'{standard.output_count_of_boxes_on_pallet}'
    # worksheet.cell(row=6, column=8).border = border

    worksheet.cell(row=7, column=1).fill = blue_fill
    worksheet.cell(row=7, column=1).value = f'Počet kusov v balení:'
    # worksheet.cell(row=7, column=1).border = border
    worksheet.cell(row=7, column=2).value = f'{standard.input_count_of_items_in_package}'
    # worksheet.cell(row=7, column=2).border = border
    worksheet.cell(row=7, column=3).fill = blue_fill
    worksheet.cell(row=7, column=3).value = f'Počet kusov na palete:'
    # worksheet.cell(row=7, column=3).border = border
    worksheet.cell(row=7, column=4).value = f'{standard.input_count_of_items_on_pallet}'
    # worksheet.cell(row=7, column=4).border = border
    worksheet.cell(row=7, column=5).fill = green_fill
    worksheet.cell(row=7, column=5).value = f'Počet kusov v balení:'
    # worksheet.cell(row=7, column=5).border = border
    worksheet.cell(row=7, column=6).value = f'{standard.output_count_of_items_in_package}'
    # worksheet.cell(row=7, column=6).border = border
    worksheet.cell(row=7, column=7).fill = green_fill
    worksheet.cell(row=7, column=7).value = f'Počet kusov na palete:'
    # worksheet.cell(row=7, column=7).border = border
    worksheet.cell(row=7, column=8).value = f'{standard.output_count_of_items_on_pallet}'
    # worksheet.cell(row=7, column=8).border = border

    worksheet.cell(row=8, column=1).fill = orange_fill
    worksheet.cell(row=8, column=1).value = f'Počet kusov na 1 pohyb:'
    # worksheet.cell(row=8, column=1).border = border
    worksheet.cell(row=8, column=2).value = f'{standard.items_per_move}'
    # worksheet.cell(row=8, column=2).border = border
    worksheet.cell(row=8, column=3).fill = orange_fill
    worksheet.cell(row=8, column=3).value = f'Jednotková váha:'
    # worksheet.cell(row=8, column=3).border = border
    worksheet.cell(row=8, column=4).value = f'{standard.unit_weight}'
    # worksheet.cell(row=8, column=4).border = border
    worksheet.cell(row=8, column=5).fill = orange_fill
    worksheet.cell(row=8, column=5).value = f'Čas prebalu:"'
    # worksheet.cell(row=8, column=5).border = border
    worksheet.cell(row=8, column=6).value = f'{standard.repacking_duration}'
    # worksheet.cell(row=8, column=6).border = border
    worksheet.cell(row=8, column=7).fill = orange_fill
    worksheet.cell(row=8, column=7).value = f'Poznámka:'
    # worksheet.cell(row=8, column=7).border = border

    worksheet.merge_cells(start_row=9, start_column=1, end_row=10, end_column=8)
    worksheet.cell(row=9, column=1).value = f'{standard.instructions}'
    # worksheet.cell(row=9, column=1).border = border

    workbook.save(response)

    return response


@permission_required('accounts.history')
@login_required
def history(request):
    cancel_sessions(request)
    repacking_history_list_all = RepackHistory.filter_and_order_repacking_history_by_get(request.GET)
    repack_history_filter = RepackHistoryFilter(request.GET, queryset=repacking_history_list_all)
    filter_GET = request.GET.copy()
    if "paginate_by" in filter_GET:
        filter_GET.pop("paginate_by")
    if "page" in filter_GET:
        filter_GET.pop("page")
    if len(filter_GET) != 0:
        filter_GET_code = "&" + filter_GET.urlencode()
    else:
        filter_GET_code = ""
    paginate_by = request.GET.get('paginate_by', 10) or 10
    open_filter = False
    if request.GET.get("paginate_by") is None and request.GET.get("page") is None and len(request.GET.keys()) != 0:
        open_filter = True
    if request.GET.get("paginate_by") is not None and request.GET.get("page") is not None and len(
            request.GET.keys()) > 2:
        open_filter = True
    if len(request.GET.keys()) > 1:
        if request.GET.get("paginate_by") is not None and request.GET.get("page") is None:
            open_filter = True
        if request.GET.get("paginate_by") is None and request.GET.get("page") is not None:
            open_filter = True

    repacking_history_list = filtered_records(request, repack_history_filter, paginate_by)
    context = {"repacking_history_list": repacking_history_list,
               'repack_history_filter': repack_history_filter, 'paginate_by': paginate_by, "open_filter": open_filter,
               "filter_GET": filter_GET_code}
    return render(request, 'repacking/history.html', context)


@permission_required('accounts.history')
@login_required
def sku_export(request):
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="SKU-export.csv"'},
    )

    response.write(u'\ufeff'.encode('utf8'))
    writer = csv.writer(response, dialect='excel', delimiter=';')
    writer.writerow(['SKU', 'COFOR', 'Destinácia', 'ks IN', 'ks OUT', 'ks v obale IN', 'ks v obale OUT',
                     'boxy IN', 'boxy OUT', 'kg/ks', 'vytvoril', 'Čas vytvorenia', 'Poznámka'])
    for standard in RepackingStandard.objects.all():
        writer.writerow([standard.SKU, standard.COFOR, standard.destination,
                         standard.input_count_of_items_on_pallet, standard.output_count_of_items_on_pallet,
                         standard.input_count_of_items_in_package, standard.output_count_of_items_in_package,
                         standard.input_count_of_boxes_on_pallet, standard.output_count_of_boxes_on_pallet,
                         standard.unit_weight, standard.creator, standard.created, standard.instructions])
    return response


@permission_required('accounts.history')
@login_required
def history_export(request):
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="history-export.csv"'},
    )

    response.write(u'\ufeff'.encode('utf8'))
    writer = csv.writer(response, dialect='excel', delimiter=';')
    RepackHistory.write_repacking_history_to_csv(RepackHistory.objects.all(), writer)
    return response


@login_required
def start(request):
    cancel_sessions(request)
    context = None

    if request.method == 'POST':
        form = RepackingForm(request.POST)
        if form.is_valid():
            if RepackingStandard.get_standard(form.cleaned_data['SKU'], form.cleaned_data['destination']) is None:
                sku = form.cleaned_data['SKU']
                destination = form.cleaned_data['destination']
                context = {"SKU": sku, 'destination': destination}
            operators = set()
            i = 1
            while f'operator_{i}' in request.POST.keys():
                operator = request.POST[f'operator_{i}']
                if operator != '':
                    try:
                        User.objects.get(barcode=operator)
                        operators.add(operator)
                    except User.DoesNotExist:
                        if context is None:
                            context = {"operator": operator}
                        else:
                            context.update({"operator": operator})
                i += 1

            if context is not None:
                return render(request, 'repacking/non_existing.html', context)
            return HttpResponseRedirect(
                f'/repacking/{form.cleaned_data["SKU"]}/{form.cleaned_data["destination"]}/{form.cleaned_data["IDP"]}/{",".join(operators)}/')

    else:
        form = RepackingForm(initial={'SKU': request.GET.get('SKU', "")})
    return render(request, 'repacking/start.html', {'form': form})


@login_required
def show_standards(request):
    cancel_sessions(request)
    repacking_standards_list_all = RepackingStandard.filter_and_order_repacking_standard_by_get(request.GET)
    standards_filter = RepackingStandardFilter(request.GET, queryset=repacking_standards_list_all)
    filter_GET = request.GET.copy()
    if "paginate_by" in filter_GET:
        filter_GET.pop("paginate_by")
    if "page" in filter_GET:
        filter_GET.pop("page")
    if len(filter_GET) != 0:
        filter_GET_code = "&" + filter_GET.urlencode()
    else:
        filter_GET_code = ""
    paginate_by = request.GET.get('paginate_by', 10) or 10
    open_filter = False
    if request.GET.get("paginate_by") is None and request.GET.get("page") is None and len(request.GET.keys()) != 0:
        open_filter = True
    if request.GET.get("paginate_by") is not None and request.GET.get("page") is not None and len(
            request.GET.keys()) > 2:
        open_filter = True
    if len(request.GET.keys()) > 1:
        if request.GET.get("paginate_by") is not None and request.GET.get("page") is None:
            open_filter = True
        if request.GET.get("paginate_by") is None and request.GET.get("page") is not None:
            open_filter = True

    repacking_standards_list = filtered_records(request, standards_filter, paginate_by)
    context = {"repacking_standards_list": repacking_standards_list,
               'standards_filter': standards_filter, 'paginate_by': paginate_by, 'open_filter': open_filter,
               'filter_GET': filter_GET_code}
    return render(request, 'repacking/standards.html', context)


@login_required
def finish(request, sku_code, destination, idp_code, operators):
    standard = RepackingStandard.get_standard(sku_code, destination)
    if standard is None:
        raise Http404("Standard does not exist")

    Log.make_log(Log.App.REPACKING, Log.Priority.DEBUG, request.user, "Prebal ukončený")

    repack_finish = datetime.now()
    repack = RepackHistory(repacking_standard=standard, idp=idp_code, repack_finish=repack_finish)
    if request.session.get(repack_start_key, False):
        repack.repack_start = request.session.get(repack_start_key)
        last_repack_start = request.session.get(repack_last_start_key)
        repack_duration = request.session.get(repack_duration_key) + (
                datetime.now() - datetime.strptime(last_repack_start,
                                                   repack_time_format)).total_seconds()
        repack.repack_duration = timedelta(seconds=repack_duration)

    else:
        Log.make_log(Log.App.REPACKING, Log.Priority.ERROR, request.user, "Prebal ukončený bez údajov oo začiatku.")

    cancel_sessions(request)

    repack.save()

    for operator in operators.split(','):
        repack.users.add(User.objects.get(barcode=operator))

    return HttpResponseRedirect('/repacking/start/')


@permission_required('accounts.sku_managment')
@login_required
def delete(request, sku_code, destination):
    standard = RepackingStandard.get_standard(sku_code, destination)
    if standard is None:
        deleted = False
    else:
        standard.delete()
        deleted = True
    return render(request, 'repacking/standard_deleted.html', {'deleted': deleted})


@permission_required('accounts.sku_managment')
@login_required
def update(request, sku_code, destination):
    # inspiracia: https://www.youtube.com/watch?v=EX6Tt-ZW0so
    standard = RepackingStandard.get_standard(sku_code, destination)
    form = StandardUpdateForm(instance=standard)
    input_photos = list(Photos.objects.all())
    for photo in input_photos:
        if photo in standard.input_photos.all():
            photo.selected = True
        else:
            photo.selected = False
    output_photos = list(Photos.objects.all())
    for photo in output_photos:
        if photo in standard.output_photos.all():
            photo.selected = True
        else:
            photo.selected = False
    tools = list(Tools.objects.all())
    for tool in tools:
        if tool in standard.tools.all():
            tool.selected = True
        else:
            tool.selected = False
    context = {'form': form, 'input_photos': input_photos, 'output_photos': output_photos, 'tools': tools}
    if request.method == 'POST':
        form = StandardUpdateForm(request.POST, instance=standard)
        if form.is_valid():
            form.save()
            standard.input_photos.clear()
            standard.output_photos.clear()
            standard.tools.clear()
            for photo_id in request.POST.getlist('existing_input_photos'):
                standard.input_photos.add(Photos.objects.get(id=photo_id))
            for photo_id in request.POST.getlist('existing_output_photos'):
                standard.output_photos.add(Photos.objects.get(id=photo_id))
            for photo_id in request.POST.getlist('existing_tools'):
                standard.tools.add(Tools.objects.get(id=photo_id))
            return redirect('/repacking/standards/')
    return render(request, 'repacking/update_standard.html', context)


@login_required
def cancel_sessions(request):
    try:
        del request.session[repack_start_key]
    except KeyError:
        pass

    try:
        del request.session[repack_last_start_key]
    except KeyError:
        pass

    try:
        del request.session[repack_duration_key]
    except KeyError:
        pass


@login_required
def cancel(request, sku_code, destination, idp_code, operators):
    cancel_sessions(request)
    return HttpResponseRedirect('/repacking/')


@login_required
def pause(request, sku_code, destination, idp_code, operators):
    repack_paused = datetime.now()
    if request.session.get(repack_duration_key, None) is not None:

        if request.session.get(repack_last_start_key, None) is not None:
            last_repack_start = request.session.get(repack_last_start_key)
            request.session[repack_last_start_key] = None
            request.session[repack_duration_key] = request.session.get(repack_duration_key) + (
                    repack_paused - datetime.strptime(last_repack_start, repack_time_format)).total_seconds()

    else:
        Log.make_log(Log.App.REPACKING, Log.Priority.ERROR, request.user, "Prebal bez začiatku bol pozastavený.")
    list_of_operator_names = []
    for operator in operators.split(','):
        list_of_operator_names.append(User.objects.get(barcode=operator).username)

    context = {'sku_code': sku_code, 'idp_code': idp_code, 'operators': operators,
               'duration': int(request.session[repack_duration_key]),
               'list_of_operator_names': list_of_operator_names}
    return render(request, 'repacking/pause.html', context)


@permission_required('accounts.sku_managment')
@login_required
def make_new_standard(request):
    cancel_sessions(request)
    if request.method == 'POST':
        form = RepackingStandardForm(request.POST, request.FILES)
        if form.is_valid() and form.cleaned_data['repacking_duration'] is not None:
            # TODO
            if RepackingStandard.get_standard(form.cleaned_data['SKU'], form.cleaned_data['destination']) is not None:
                raise FileExistsError("RepackingForm standard w/ this SKU already exists")

            standard = RepackingStandard(
                SKU=form.cleaned_data['SKU'],
                COFOR=form.cleaned_data['COFOR'],
                supplier=form.cleaned_data['supplier'],
                destination=form.cleaned_data['destination'],
                items_per_move=form.cleaned_data['items_per_move'],
                unit_weight=form.cleaned_data['unit_weight'],
                repacking_duration=timedelta(seconds=form.cleaned_data['repacking_duration'].total_seconds()),
                instructions=form.cleaned_data['instructions'],
                input_count_of_items_in_package=form.cleaned_data['input_count_of_items_in_package'],
                output_count_of_items_in_package=form.cleaned_data['output_count_of_items_in_package'],
                input_count_of_boxes_on_pallet=form.cleaned_data['input_count_of_boxes_on_pallet'],
                output_count_of_boxes_on_pallet=form.cleaned_data['output_count_of_boxes_on_pallet'],
                input_count_of_items_on_pallet=form.cleaned_data['input_count_of_items_on_pallet'],
                output_count_of_items_on_pallet=form.cleaned_data['output_count_of_items_on_pallet'],
                input_type_of_package=form.cleaned_data['input_type_of_package'],
                output_type_of_package=form.cleaned_data['output_type_of_package'],
                creator=request.user
            )
            standard.save()

            print(request.POST.getlist('existing_input_photos'))
            print(request.POST.getlist('existing_output_photos'))
            print(request.POST.getlist('existing_tools'))
            for photo_id in request.POST.getlist('existing_input_photos'):
                standard.input_photos.add(Photos.objects.get(id=photo_id))
            for photo_id in request.POST.getlist('existing_output_photos'):
                standard.output_photos.add(Photos.objects.get(id=photo_id))
            for photo_id in request.POST.getlist('existing_tools'):
                standard.tools.add(Tools.objects.get(id=photo_id))

            if 'input_photos' in form.files:
                for photo in form.files.getlist('input_photos'):
                    input_photo = Photos(photo=photo)
                    input_photo.save()
                    standard.input_photos.add(input_photo)
            if 'output_photos' in form.files:
                for photo in form.files.getlist('output_photos'):
                    output_photo = Photos(photo=photo)
                    output_photo.save()
                    standard.output_photos.add(output_photo)
            if 'tools' in form.files:
                for photo in form.files.getlist('tools'):
                    tool = Tools(photo=photo)
                    tool.save()
                    standard.tools.add(tool)

            Log.make_log(Log.App.REPACKING, Log.Priority.DEBUG, request.user, "Vytvorený nový štandard")

            return HttpResponseRedirect("/")

    else:
        form = RepackingStandardForm()

    return render(request, 'repacking/new_standard.html',
                  {'form': form, 'photos': Photos.objects.all(), 'tools': Tools.objects.all()})
